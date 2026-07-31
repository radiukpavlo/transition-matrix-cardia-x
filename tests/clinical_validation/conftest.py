from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from tm_ecg.clinical_validation.workbook_reader import (
    B2_FIELDS,
    CASE_SHEET_CELLS,
    EXPECTED_RESPONSE_HEADERS,
)


LABEL_DIAGNOSES = {
    "Normal": "Normal sinus rhythm",
    "AF": "Atrial fibrillation",
    "PVC": "Premature ventricular complexes",
    "RBBB spectrum": "Right bundle branch block",
    "ST depression": "ST depression",
}


def _b1_label(index: int) -> str:
    labels = tuple(LABEL_DIAGNOSES)
    return labels[((index - 1) // 20) % len(labels)]


@pytest.fixture()
def synthetic_validation_inputs(tmp_path: Path) -> tuple[Path, Path]:
    workbook_path = tmp_path / "completed.xlsx"
    provenance_path = tmp_path / "case_provenance.json"
    workbook = Workbook()
    workbook.remove(workbook.active)
    case_index = workbook.create_sheet("Case_Index")
    case_index.append(["case_id"])
    response_export = workbook.create_sheet("Response_Export")
    response_export.append(list(EXPECTED_RESPONSE_HEADERS))
    rule_review = workbook.create_sheet("Rule_Review")
    rule_review.append([])
    rule_review.append([])
    rule_review.append([])
    rule_review.append([])
    rule_review.append(
        [
            "scope",
            "rule_id",
            "antecedent combination",
            "consequent label",
            "A1 support",
            "A2 conservative",
            "A3 understandable",
            "A4 exclusions missing?",
            "A4 exclusion notes",
            "A5 useful",
        ]
    )
    for index in range(10):
        rule_review.append(["B1" if index < 8 else "B2", "", "", "", "", "", "", "", "", ""])
    provenance: dict[str, dict[str, object]] = {}

    cases: list[tuple[str, str, str, str, str]] = []
    for index in range(1, 101):
        case_id = f"P{index:03d}"
        route = "exact_match" if index <= 30 else "conflict_region" if index <= 60 else "structural_abstention"
        label = _b1_label(index)
        cases.append((case_id, "B1", route, label, ""))
    for index in range(1, 11):
        case_id = f"D{index:03d}"
        original = f"P{index:03d}"
        route = "exact_match"
        label = _b1_label(index)
        cases.append((case_id, "B1", route, label, original))
    for index in range(1, 41):
        cases.append((f"L{index:03d}", "B2", "morphology_audit", "", ""))

    for ordinal, (case_id, branch, route, label, duplicate_of) in enumerate(cases, start=1):
        diagnosis = LABEL_DIAGNOSES.get(label, "")
        row = {header: "" for header in EXPECTED_RESPONSE_HEADERS}
        row.update(
            {
                "case_id": case_id,
                "dataset_branch": branch,
                "display_id": case_id,
                "cardia_x_route": route,
                "primary_diagnosis": diagnosis,
                "diagnostic_confidence": 5 if branch == "B1" else "",
                "ecg_quality": 5 if branch == "B1" else "",
                "ambiguous": "No" if branch == "B1" else "",
                "requires_additional_information": "No" if branch == "B1" else "",
                "rationale": diagnosis,
            }
        )
        if branch == "B2":
            for field in B2_FIELDS[:5]:
                row[field] = 4
        response_export.append([row[header] for header in EXPECTED_RESPONSE_HEADERS])
        case_index.append([case_id])
        sheet = workbook.create_sheet(case_id)
        sheet["M5"] = case_id
        for field, coordinate in CASE_SHEET_CELLS.items():
            sheet[coordinate] = row[field]
        provenance[case_id] = {
            "record_id": f"synthetic-{ordinal}",
            "route_category": route,
            "hidden_benchmark_label": label,
            "duplicate_of_case_id": duplicate_of,
            "duplicate_status": "repeat" if duplicate_of else "original",
        }

    workbook.save(workbook_path)
    provenance_path.write_text(
        json.dumps({"case_provenance": provenance}, indent=2), encoding="utf-8"
    )
    return workbook_path, provenance_path

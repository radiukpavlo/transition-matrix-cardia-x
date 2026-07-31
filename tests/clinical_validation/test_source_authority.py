from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
import pytest

from tm_ecg.clinical_validation.audit import sha256_file
from tm_ecg.clinical_validation.source_authority import (
    SourceAuthorityError,
    allows_unsigned_source_exploration,
    hash_normalized_responses,
    load_response_authority,
    normalize_response_export,
    resolve_unassisted_source,
    verify_workbook_hash,
    write_authority_audit,
)
from tm_ecg.clinical_validation.workbook_reader import (
    CASE_SHEET_CELLS,
    WorkbookSchemaError,
    load_provenance,
    read_completed_workbook,
)


CANONICAL_FIELDS = tuple(CASE_SHEET_CELLS)
ROOT = Path(__file__).resolve().parents[2]
COMMITTED_WORKBOOK = (
    ROOT / "clinical_validation" / "review" / "CARDIA-X_clinical_validation_v6.xlsx"
)
COMMITTED_PROVENANCE = (
    ROOT / "clinical_validation" / "cases" / "private" / "case_provenance.json"
)
requires_private_reader_evidence = pytest.mark.skipif(
    not (COMMITTED_WORKBOOK.is_file() and COMMITTED_PROVENANCE.is_file()),
    reason="case-level reader evidence is intentionally excluded from the public repository",
)


def _write_manifest(
    path: Path,
    workbook: Path,
    *,
    normalized_hash: str,
    expected_conflicts: int,
    signed: bool,
) -> Path:
    payload = {
        "version": 1,
        "workbook": {
            "path": str(workbook),
            "sha256": sha256_file(workbook),
            "immutable": True,
        },
        "canonical_unassisted_response_source": {
            "sheet": "Response_Export",
            "normalized_payload_sha256": normalized_hash,
            "fields": list(CANONICAL_FIELDS),
        },
        "non_authoritative_mirrors": {
            "per_case_sheets": {
                "status": "deterministic_benchmark_stress_test",
                "may_supply_physician_response": False,
            }
        },
        "conflict_policy": {
            "edit_workbook": False,
            "fail_on_unknown_conflict": True,
            "emit_complete_conflict_ledger": True,
            "expected_mirror_conflict_count": expected_conflicts,
        },
        "expected_population": {
            "canonical_row_count": 150,
            "unique_b1_case_count": 100,
        },
        "signoff": {
            "decision_id": "TEST-SIGNED-001" if signed else "pending_project_owner_signoff",
            "signed_by": "test-owner" if signed else None,
            "signed_at_utc": "2026-07-24T00:00:00Z" if signed else None,
        },
        "provenance_discrepancy": {},
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _original_case_ids(provenance_path: Path) -> dict[str, str]:
    return {
        case_id: str(row.get("duplicate_of_case_id", "")).strip() or case_id
        for case_id, row in load_provenance(provenance_path).items()
    }


def test_normalized_response_hash_is_deterministic(synthetic_validation_inputs) -> None:
    workbook, _ = synthetic_validation_inputs
    first = normalize_response_export(workbook, fields=CANONICAL_FIELDS)
    second = normalize_response_export(workbook, fields=CANONICAL_FIELDS)
    assert first == second
    assert hash_normalized_responses(first) == hash_normalized_responses(second)
    assert len(first) == 150
    assert all(row["row_sha256"] for row in first)


def test_hash_verification_rejects_one_byte_mutation(
    synthetic_validation_inputs,
) -> None:
    workbook, _ = synthetic_validation_inputs
    expected = sha256_file(workbook)
    with workbook.open("ab") as handle:
        handle.write(b"x")
    with pytest.raises(SourceAuthorityError, match="SHA-256 mismatch"):
        verify_workbook_hash(workbook, expected)


def test_authority_mode_never_uses_case_sheet_response(
    synthetic_validation_inputs,
    tmp_path: Path,
) -> None:
    workbook_path, provenance = synthetic_validation_inputs
    workbook = load_workbook(workbook_path)
    canonical_diagnosis = workbook["Response_Export"]["O2"].value
    workbook["P001"]["M7"] = "tampered benchmark-derived mirror"
    workbook.save(workbook_path)

    rows = normalize_response_export(
        workbook_path,
        fields=CANONICAL_FIELDS,
        original_case_ids=_original_case_ids(provenance),
    )
    manifest_path = _write_manifest(
        tmp_path / "authority.json",
        workbook_path,
        normalized_hash=hash_normalized_responses(rows),
        expected_conflicts=1,
        signed=True,
    )
    output_dir = tmp_path / "audit"
    extraction = read_completed_workbook(
        workbook_path,
        provenance,
        reconciliation_mode="authority_manifest",
        response_authority_path=manifest_path,
        authority_output_dir=output_dir,
    )
    assert extraction.physician_responses[0].primary_diagnosis == canonical_diagnosis
    assert (
        dict(extraction.physician_responses[0].source_cells)["primary_diagnosis"]
        == "Response_Export!O2"
    )
    assert extraction.response_import_audit["source_authority_gate"]["passed"] is True
    assert (output_dir / "workbook_conflict_ledger.csv").exists()


def test_unsigned_manifest_blocks_authority_reader(
    synthetic_validation_inputs,
    tmp_path: Path,
) -> None:
    workbook, provenance = synthetic_validation_inputs
    rows = normalize_response_export(workbook, fields=CANONICAL_FIELDS)
    manifest_path = _write_manifest(
        tmp_path / "authority.json",
        workbook,
        normalized_hash=hash_normalized_responses(rows),
        expected_conflicts=0,
        signed=False,
    )
    with pytest.raises(WorkbookSchemaError, match="manifest_signed"):
        read_completed_workbook(
            workbook,
            provenance,
            reconciliation_mode="authority_manifest",
            response_authority_path=manifest_path,
        )


def test_unsigned_manifest_allows_explicit_exploratory_read_only(
    synthetic_validation_inputs,
    tmp_path: Path,
) -> None:
    workbook, provenance = synthetic_validation_inputs
    rows = normalize_response_export(workbook, fields=CANONICAL_FIELDS)
    manifest_path = _write_manifest(
        tmp_path / "authority.json",
        workbook,
        normalized_hash=hash_normalized_responses(rows),
        expected_conflicts=0,
        signed=False,
    )
    before_hash = sha256_file(workbook)
    extraction = read_completed_workbook(
        workbook,
        provenance,
        reconciliation_mode="authority_manifest",
        response_authority_path=manifest_path,
        allow_unsigned_source_exploration=True,
    )

    gate = extraction.response_import_audit["source_authority_gate"]
    assert allows_unsigned_source_exploration(gate) is True
    assert gate["passed"] is False
    assert gate["checks"]["manifest_signed"] is False
    assert len(extraction.physician_responses) == 150
    assert sha256_file(workbook) == before_hash


def test_unsigned_exploration_rejects_any_second_integrity_failure() -> None:
    gate = {
        "passed": False,
        "source_authority_status": "source_authority_unresolved",
        "checks": {
            "manifest_signed": False,
            "workbook_sha256_match": False,
            "workbook_unchanged_during_read": True,
        },
    }
    assert allows_unsigned_source_exploration(gate) is False


def test_missing_canonical_field_fails_closed(
    synthetic_validation_inputs,
    tmp_path: Path,
) -> None:
    workbook_path, _ = synthetic_validation_inputs
    workbook = load_workbook(workbook_path)
    workbook["Response_Export"]["O1"] = "renamed_primary_diagnosis"
    workbook.save(workbook_path)
    manifest_path = _write_manifest(
        tmp_path / "authority.json",
        workbook_path,
        normalized_hash="0" * 64,
        expected_conflicts=0,
        signed=True,
    )
    manifest = load_response_authority(manifest_path)
    with pytest.raises(SourceAuthorityError, match="missing canonical fields"):
        resolve_unassisted_source(workbook_path, manifest)


@requires_private_reader_evidence
def test_committed_v6_audit_emits_all_known_conflicts(tmp_path: Path) -> None:
    workbook = COMMITTED_WORKBOOK
    provenance = COMMITTED_PROVENANCE
    manifest = load_response_authority(
        ROOT / "clinical_validation" / "config" / "response_authority_v1.yaml"
    )
    before_hash = sha256_file(workbook)
    resolution = resolve_unassisted_source(
        workbook,
        manifest,
        original_case_ids=_original_case_ids(provenance),
    )
    paths = write_authority_audit(tmp_path, resolution)

    assert len(resolution.normalized_rows) == 150
    assert len(resolution.conflict_ledger) == 973
    assert resolution.consistency_audit["cases_with_mirror_conflicts"] == 150
    assert resolution.consistency_audit["unknown_conflict_count"] == 0
    assert resolution.gate["checks"]["workbook_sha256_match"] is True
    assert resolution.gate["checks"]["normalized_response_sha256_match"] is True
    assert resolution.gate["checks"]["manifest_signed"] is False
    assert resolution.gate["status"] == "fail"
    assert sha256_file(workbook) == before_hash
    assert set(paths) == {
        "manifest",
        "consistency_audit",
        "conflict_ledger",
        "normalized_responses",
        "normalized_responses_sha256",
        "gate",
    }


@requires_private_reader_evidence
def test_committed_v6_exploratory_reader_never_backfills_b2_stress_test_cells(
    tmp_path: Path,
) -> None:
    workbook = COMMITTED_WORKBOOK
    provenance = COMMITTED_PROVENANCE
    manifest = ROOT / "clinical_validation" / "config" / "response_authority_v1.yaml"
    before_hash = sha256_file(workbook)

    extraction = read_completed_workbook(
        workbook,
        provenance,
        reconciliation_mode="authority_manifest",
        response_authority_path=manifest,
        authority_output_dir=tmp_path,
        allow_unsigned_source_exploration=True,
    )

    assert extraction.response_import_audit["response_export_schema"] == ("response_export_v6")
    assert extraction.response_import_audit["b2_morphology_source"] == (
        "Response_Export_partial; structured B2 ratings unavailable"
    )
    assert extraction.response_import_audit["b2_missing_response_export_fields"] == [
        "p_wave_delineation_plausible",
        "qrs_delineation_plausible",
        "t_wave_delineation_plausible",
        "lead_quality_reasonable",
        "lead_quality_utility",
    ]
    assert len(extraction.b2_responses) == 40
    assert all(
        row["p_wave_delineation_plausible"] is None
        and row["qrs_delineation_plausible"] is None
        and row["t_wave_delineation_plausible"] is None
        and row["lead_quality_reasonable"] is None
        and row["lead_quality_utility"] is None
        for row in extraction.b2_responses
    )
    assert extraction.b2_responses[0]["morphology_comment"] == "normal morphology"
    assert sha256_file(workbook) == before_hash

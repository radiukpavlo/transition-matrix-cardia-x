"""Detached, fail-closed source authority for immutable validation workbooks."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
import re
from typing import Any, Mapping, Sequence
import unicodedata

from tm_ecg.clinical_validation.audit import (
    sha256_file,
    sha256_payload,
    write_csv,
    write_json,
    write_jsonl,
)
from tm_ecg.clinical_validation.ontology import load_json_yaml


SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")


class SourceAuthorityError(ValueError):
    """Raised when a detached source-authority contract is invalid."""


def allows_unsigned_source_exploration(gate: Mapping[str, object]) -> bool:
    """Return true only when owner signoff is the sole failed G0 check."""

    checks = gate.get("checks")
    if not isinstance(checks, Mapping):
        return False
    if checks.get("manifest_signed") is not False:
        return False
    non_signoff_checks = [value for name, value in checks.items() if str(name) != "manifest_signed"]
    return (
        gate.get("passed") is False
        and gate.get("source_authority_status") == "source_authority_unresolved"
        and bool(non_signoff_checks)
        and all(value is True for value in non_signoff_checks)
    )


@dataclass(frozen=True, slots=True)
class ResponseAuthorityManifest:
    version: int
    workbook_path: str
    workbook_sha256: str
    workbook_immutable: bool
    canonical_sheet: str
    normalized_payload_sha256: str
    fields: tuple[str, ...]
    mirror_status: str
    mirror_may_supply_physician_response: bool
    edit_workbook: bool
    fail_on_unknown_conflict: bool
    emit_complete_conflict_ledger: bool
    expected_mirror_conflict_count: int
    expected_canonical_row_count: int
    expected_unique_b1_case_count: int
    decision_id: str
    signed_by: str | None
    signed_at_utc: str | None
    plan_registered_sha256: str | None = None
    provenance_note: str | None = None

    @property
    def is_signed(self) -> bool:
        return bool(
            self.decision_id
            and not self.decision_id.lower().startswith("pending")
            and self.signed_by
            and self.signed_at_utc
        )

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class SourceAuthorityResolution:
    manifest: ResponseAuthorityManifest
    normalized_rows: tuple[dict[str, Any], ...]
    conflict_ledger: tuple[dict[str, Any], ...]
    consistency_audit: dict[str, Any]
    gate: dict[str, Any]


def _object(
    payload: Mapping[str, object],
    field: str,
    *,
    location: str,
) -> Mapping[str, object]:
    value = payload.get(field)
    if not isinstance(value, Mapping):
        raise SourceAuthorityError(f"{location}.{field} must be an object")
    return value


def _string(
    payload: Mapping[str, object],
    field: str,
    *,
    location: str,
    allow_empty: bool = False,
) -> str:
    value = payload.get(field)
    if not isinstance(value, str) or (not allow_empty and not value.strip()):
        raise SourceAuthorityError(f"{location}.{field} must be a non-empty string")
    return value.strip()


def _boolean(payload: Mapping[str, object], field: str, *, location: str) -> bool:
    value = payload.get(field)
    if not isinstance(value, bool):
        raise SourceAuthorityError(f"{location}.{field} must be a boolean")
    return value


def _integer(
    payload: Mapping[str, object],
    field: str,
    *,
    location: str,
    minimum: int = 0,
) -> int:
    value = payload.get(field)
    if isinstance(value, bool) or not isinstance(value, int) or value < minimum:
        raise SourceAuthorityError(
            f"{location}.{field} must be an integer greater than or equal to {minimum}"
        )
    return value


def _sha256(payload: Mapping[str, object], field: str, *, location: str) -> str:
    value = _string(payload, field, location=location).lower()
    if not SHA256_PATTERN.fullmatch(value):
        raise SourceAuthorityError(f"{location}.{field} must be a lowercase SHA-256")
    return value


def load_response_authority(path: str | Path) -> ResponseAuthorityManifest:
    """Load and strictly validate a detached response-authority manifest."""

    payload = load_json_yaml(path)
    allowed_root = {
        "version",
        "workbook",
        "canonical_unassisted_response_source",
        "non_authoritative_mirrors",
        "conflict_policy",
        "expected_population",
        "signoff",
        "provenance_discrepancy",
    }
    unknown = sorted(set(payload) - allowed_root)
    if unknown:
        raise SourceAuthorityError(f"Unknown response-authority fields: {unknown}")
    if payload.get("version") != 1:
        raise SourceAuthorityError("Response authority must declare version 1")

    workbook = _object(payload, "workbook", location="manifest")
    canonical = _object(payload, "canonical_unassisted_response_source", location="manifest")
    mirrors = _object(payload, "non_authoritative_mirrors", location="manifest")
    case_sheets = _object(mirrors, "per_case_sheets", location="non_authoritative_mirrors")
    conflict = _object(payload, "conflict_policy", location="manifest")
    population = _object(payload, "expected_population", location="manifest")
    signoff = _object(payload, "signoff", location="manifest")
    discrepancy_value = payload.get("provenance_discrepancy", {})
    if not isinstance(discrepancy_value, Mapping):
        raise SourceAuthorityError("manifest.provenance_discrepancy must be an object")
    discrepancy = discrepancy_value

    raw_fields = canonical.get("fields")
    if not isinstance(raw_fields, list) or not raw_fields:
        raise SourceAuthorityError(
            "canonical_unassisted_response_source.fields must be a non-empty list"
        )
    fields = tuple(str(item).strip() for item in raw_fields)
    if any(not field for field in fields) or len(set(fields)) != len(fields):
        raise SourceAuthorityError("Canonical response fields must be non-empty and unique")

    signed_by_value = signoff.get("signed_by")
    signed_at_value = signoff.get("signed_at_utc")
    signed_by = str(signed_by_value).strip() if signed_by_value not in {None, ""} else None
    signed_at = str(signed_at_value).strip() if signed_at_value not in {None, ""} else None
    plan_hash = discrepancy.get("plan_registered_sha256")
    if plan_hash not in {None, ""}:
        if not isinstance(plan_hash, str) or not SHA256_PATTERN.fullmatch(plan_hash.lower()):
            raise SourceAuthorityError(
                "provenance_discrepancy.plan_registered_sha256 must be a SHA-256"
            )
        plan_hash = plan_hash.lower()
    note_value = discrepancy.get("note")

    manifest = ResponseAuthorityManifest(
        version=1,
        workbook_path=_string(workbook, "path", location="workbook"),
        workbook_sha256=_sha256(workbook, "sha256", location="workbook"),
        workbook_immutable=_boolean(workbook, "immutable", location="workbook"),
        canonical_sheet=_string(canonical, "sheet", location="canonical_source"),
        normalized_payload_sha256=_sha256(
            canonical, "normalized_payload_sha256", location="canonical_source"
        ),
        fields=fields,
        mirror_status=_string(case_sheets, "status", location="per_case_sheets"),
        mirror_may_supply_physician_response=_boolean(
            case_sheets, "may_supply_physician_response", location="per_case_sheets"
        ),
        edit_workbook=_boolean(conflict, "edit_workbook", location="conflict_policy"),
        fail_on_unknown_conflict=_boolean(
            conflict, "fail_on_unknown_conflict", location="conflict_policy"
        ),
        emit_complete_conflict_ledger=_boolean(
            conflict, "emit_complete_conflict_ledger", location="conflict_policy"
        ),
        expected_mirror_conflict_count=_integer(
            conflict, "expected_mirror_conflict_count", location="conflict_policy"
        ),
        expected_canonical_row_count=_integer(
            population, "canonical_row_count", location="expected_population", minimum=1
        ),
        expected_unique_b1_case_count=_integer(
            population, "unique_b1_case_count", location="expected_population", minimum=1
        ),
        decision_id=_string(signoff, "decision_id", location="signoff"),
        signed_by=signed_by,
        signed_at_utc=signed_at,
        plan_registered_sha256=plan_hash,
        provenance_note=(str(note_value).strip() if note_value not in {None, ""} else None),
    )
    if not manifest.workbook_immutable:
        raise SourceAuthorityError("The completed workbook must be declared immutable")
    if manifest.edit_workbook:
        raise SourceAuthorityError("The conflict policy must prohibit workbook edits")
    if manifest.mirror_may_supply_physician_response:
        raise SourceAuthorityError(
            "Non-authoritative case sheets may not supply physician responses"
        )
    if manifest.canonical_sheet != "Response_Export":
        raise SourceAuthorityError("Only the explicitly reviewed Response_Export is supported")
    return manifest


def verify_workbook_hash(
    workbook: str | Path,
    expected_sha256: str,
    *,
    raise_on_mismatch: bool = True,
) -> str:
    """Return the workbook hash and optionally reject a mismatch."""

    actual = sha256_file(workbook)
    expected = expected_sha256.lower()
    if actual != expected and raise_on_mismatch:
        raise SourceAuthorityError(
            f"Workbook SHA-256 mismatch: expected {expected}, observed {actual}"
        )
    return actual


def _normalized_cell(value: object) -> tuple[str, tuple[str, ...]]:
    operations: list[str] = []
    if value is None:
        return "", ("null_to_empty",)
    if isinstance(value, bool):
        return ("true" if value else "false"), ("boolean_to_lowercase_text",)
    if isinstance(value, int):
        return str(value), ("integer_to_text",)
    if isinstance(value, float):
        normalized_number = str(int(value)) if value.is_integer() else format(value, ".15g")
        return normalized_number, ("number_to_canonical_text",)
    raw = str(value)
    normalized = unicodedata.normalize("NFKC", raw)
    if normalized != raw:
        operations.append("unicode_nfkc")
    stripped = normalized.strip()
    if stripped != normalized:
        operations.append("trim")
    collapsed = re.sub(r"\s+", " ", stripped)
    if collapsed != stripped:
        operations.append("collapse_whitespace")
    return collapsed, tuple(operations)


def _raw_json_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _load_read_only_workbook(workbook: str | Path) -> object:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Clinical validation requires openpyxl>=3.1") from exc
    return load_workbook(Path(workbook), read_only=True, data_only=True)


def normalize_response_export(
    workbook: str | Path | object,
    *,
    fields: Sequence[str],
    original_case_ids: Mapping[str, str] | None = None,
) -> tuple[dict[str, Any], ...]:
    """Normalize only the canonical response fields with full cell provenance."""

    opened_here = isinstance(workbook, (str, Path))
    if isinstance(workbook, (str, Path)):
        book = _load_read_only_workbook(workbook)
    else:
        book = workbook
    try:
        if "Response_Export" not in book.sheetnames:  # type: ignore[attr-defined]
            raise SourceAuthorityError("Workbook is missing Response_Export")
        sheet = book["Response_Export"]  # type: ignore[index]
        values = sheet.iter_rows(values_only=True)  # type: ignore[attr-defined]
        try:
            header_values = next(values)
        except StopIteration as exc:
            raise SourceAuthorityError("Response_Export is empty") from exc
        headers = [_normalized_cell(value)[0] for value in header_values]
        if len(set(headers)) != len(headers):
            raise SourceAuthorityError("Response_Export contains duplicate headers")
        missing = sorted({"case_id", "dataset_branch", *fields} - set(headers))
        if missing:
            raise SourceAuthorityError(f"Response_Export is missing canonical fields: {missing}")
        column_by_field = {
            field: headers.index(field) + 1 for field in ("case_id", "dataset_branch", *fields)
        }
        try:
            from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("Clinical validation requires openpyxl>=3.1") from exc

        normalized_rows: list[dict[str, Any]] = []
        seen_case_ids: set[str] = set()
        for row_number, row_values in enumerate(values, start=2):
            if not any(value not in {None, ""} for value in row_values):
                continue
            padded = tuple(row_values) + (None,) * max(0, len(headers) - len(row_values))
            row = dict(zip(headers, padded, strict=False))
            case_id, case_operations = _normalized_cell(row.get("case_id"))
            if not case_id:
                raise SourceAuthorityError(f"Response_Export row {row_number} has no case_id")
            if case_id in seen_case_ids:
                raise SourceAuthorityError(f"Duplicate Response_Export case_id {case_id}")
            seen_case_ids.add(case_id)
            dataset_branch, branch_operations = _normalized_cell(row.get("dataset_branch"))
            field_rows: dict[str, dict[str, Any]] = {}
            source_cells: dict[str, str] = {}
            normalized_values: dict[str, str] = {}
            for field in fields:
                normalized_value, operations = _normalized_cell(row.get(field))
                column = get_column_letter(column_by_field[field])
                coordinate = f"Response_Export!{column}{row_number}"
                source_cells[field] = coordinate
                normalized_values[field] = normalized_value
                field_rows[field] = {
                    "raw_value": _raw_json_value(row.get(field)),
                    "normalized_value": normalized_value,
                    "normalization_operations": list(operations),
                    "source_cell": coordinate,
                }
            original_case_id = (
                original_case_ids.get(case_id, case_id)
                if original_case_ids is not None
                else case_id
            )
            content = {
                "workbook_case_id": case_id,
                "original_case_id": original_case_id,
                "dataset_branch": dataset_branch,
                "source_sheet": "Response_Export",
                "source_row": row_number,
                "case_id_normalization_operations": list(case_operations),
                "dataset_branch_normalization_operations": list(branch_operations),
                "source_cells": source_cells,
                "normalized_values": normalized_values,
                "fields": field_rows,
            }
            normalized_rows.append({**content, "row_sha256": sha256_payload(content)})
        return tuple(normalized_rows)
    finally:
        if opened_here:
            book.close()  # type: ignore[attr-defined]


def _normalized_content(rows: Sequence[Mapping[str, object]]) -> dict[str, object]:
    content_rows: list[dict[str, object]] = []
    for row in rows:
        values = row.get("normalized_values")
        if not isinstance(values, Mapping):
            raise SourceAuthorityError("Normalized response row has no normalized_values")
        content_rows.append(
            {
                "workbook_case_id": str(row.get("workbook_case_id", "")),
                "values": {str(key): str(value) for key, value in values.items()},
            }
        )
    return {"normalization_contract": "response_export_nfkc_v1", "rows": content_rows}


def hash_normalized_responses(rows: Sequence[Mapping[str, object]]) -> str:
    """Hash normalized values, excluding non-content provenance and row hashes."""

    return sha256_payload(_normalized_content(rows))


def _case_sheet_cells() -> Mapping[str, str]:
    # Local import avoids a module cycle: workbook_reader imports this module.
    from tm_ecg.clinical_validation.workbook_reader import CASE_SHEET_CELLS

    return CASE_SHEET_CELLS


def compare_all_mirrors(
    workbook: str | Path | object,
    rows: Sequence[Mapping[str, object]],
    *,
    fields: Sequence[str],
    mirror_status: str,
) -> tuple[tuple[dict[str, Any], ...], dict[str, Any]]:
    """Compare every declared mirror without ever using it as a response source."""

    opened_here = isinstance(workbook, (str, Path))
    if isinstance(workbook, (str, Path)):
        book = _load_read_only_workbook(workbook)
    else:
        book = workbook
    ledger: list[dict[str, Any]] = []
    mismatch_by_field = {field: 0 for field in fields}
    unknown_conflicts: list[dict[str, str]] = []
    try:
        case_cells = _case_sheet_cells()
        missing_mappings = sorted(set(fields) - set(case_cells))
        if missing_mappings:
            raise SourceAuthorityError(
                f"Canonical fields have no declared case-sheet mirror: {missing_mappings}"
            )
        for normalized_row in rows:
            case_id = str(normalized_row["workbook_case_id"])
            if case_id not in book.sheetnames:  # type: ignore[attr-defined]
                unknown_conflicts.append({"case_id": case_id, "reason": "missing_case_sheet"})
                continue
            sheet = book[case_id]  # type: ignore[index]
            if _normalized_cell(sheet["M5"].value)[0] != case_id:
                unknown_conflicts.append(
                    {"case_id": case_id, "reason": "case_sheet_identity_mismatch"}
                )
                continue
            canonical_fields = normalized_row.get("fields")
            if not isinstance(canonical_fields, Mapping):
                raise SourceAuthorityError(f"Normalized response {case_id} has no field provenance")
            for field in fields:
                canonical_entry = canonical_fields.get(field)
                if not isinstance(canonical_entry, Mapping):
                    raise SourceAuthorityError(
                        f"Normalized response {case_id} is missing field {field}"
                    )
                mirror_coordinate = case_cells[field]
                mirror_raw = sheet[mirror_coordinate].value
                mirror_value, mirror_operations = _normalized_cell(mirror_raw)
                canonical_value = str(canonical_entry.get("normalized_value", ""))
                if mirror_value == canonical_value:
                    continue
                mismatch_by_field[field] += 1
                ledger.append(
                    {
                        "workbook_case_id": case_id,
                        "original_case_id": str(normalized_row.get("original_case_id", case_id)),
                        "field": field,
                        "canonical_source": "Response_Export",
                        "canonical_cell": str(canonical_entry.get("source_cell", "")),
                        "canonical_raw_value": canonical_entry.get("raw_value"),
                        "canonical_normalized_value": canonical_value,
                        "mirror_source": "per_case_sheet",
                        "mirror_cell": f"{case_id}!{mirror_coordinate}",
                        "mirror_raw_value": _raw_json_value(mirror_raw),
                        "mirror_normalized_value": mirror_value,
                        "mirror_normalization_operations": list(mirror_operations),
                        "classification": mirror_status,
                        "canonical_value_selected": True,
                    }
                )
        audit = {
            "canonical_row_count": len(rows),
            "mirror_conflict_count": len(ledger),
            "mirror_conflicts_by_field": mismatch_by_field,
            "cases_with_mirror_conflicts": len({str(row["workbook_case_id"]) for row in ledger}),
            "unknown_conflict_count": len(unknown_conflicts),
            "unknown_conflicts": unknown_conflicts,
            "case_sheet_values_reconciled": len(ledger) == 0,
            "case_sheets_used_as_response_source": False,
            "mirror_status": mirror_status,
        }
        return tuple(ledger), audit
    finally:
        if opened_here:
            book.close()  # type: ignore[attr-defined]


def resolve_unassisted_source(
    workbook: str | Path,
    manifest: ResponseAuthorityManifest,
    *,
    original_case_ids: Mapping[str, str] | None = None,
) -> SourceAuthorityResolution:
    """Resolve the declared canonical table and produce a complete gate audit."""

    workbook_path = Path(workbook)
    before_hash = verify_workbook_hash(
        workbook_path, manifest.workbook_sha256, raise_on_mismatch=False
    )
    workbook_hash_match = before_hash == manifest.workbook_sha256
    rows = normalize_response_export(
        workbook_path,
        fields=manifest.fields,
        original_case_ids=original_case_ids,
    )
    normalized_hash = hash_normalized_responses(rows)
    ledger, audit = compare_all_mirrors(
        workbook_path,
        rows,
        fields=manifest.fields,
        mirror_status=manifest.mirror_status,
    )
    after_hash = sha256_file(workbook_path)
    b1_rows = [
        row
        for row in rows
        if str(row.get("dataset_branch", "")) == "B1"
        or str(row["workbook_case_id"]).startswith("CX-B1-")
    ]
    unique_b1_count = len(
        {str(row.get("original_case_id", row["workbook_case_id"])) for row in b1_rows}
    )
    gate_checks = {
        "workbook_sha256_match": workbook_hash_match,
        "normalized_response_sha256_match": (normalized_hash == manifest.normalized_payload_sha256),
        "canonical_row_count_match": (len(rows) == manifest.expected_canonical_row_count),
        "unique_b1_case_count_match": (unique_b1_count == manifest.expected_unique_b1_case_count),
        "mirror_conflict_count_match": (len(ledger) == manifest.expected_mirror_conflict_count),
        "unknown_conflict_count_zero": audit["unknown_conflict_count"] == 0,
        "manifest_signed": manifest.is_signed,
        "workbook_unchanged_during_read": before_hash == after_hash,
        "case_sheets_not_used_as_response_source": not audit["case_sheets_used_as_response_source"],
    }
    gate = {
        "gate_id": "G0_source_authority",
        "status": "pass" if all(gate_checks.values()) else "fail",
        "passed": all(gate_checks.values()),
        "checks": gate_checks,
        "workbook_expected_sha256": manifest.workbook_sha256,
        "workbook_observed_sha256": before_hash,
        "workbook_post_read_sha256": after_hash,
        "normalized_response_expected_sha256": manifest.normalized_payload_sha256,
        "normalized_response_observed_sha256": normalized_hash,
        "canonical_row_count": len(rows),
        "unique_b1_case_count": unique_b1_count,
        "mirror_conflict_count": len(ledger),
        "unknown_conflict_count": audit["unknown_conflict_count"],
        "workbook_write_attempts": 0,
        "source_authority_status": (
            "resolved" if manifest.is_signed else "source_authority_unresolved"
        ),
    }
    consistency = {
        **audit,
        "workbook_expected_sha256": manifest.workbook_sha256,
        "workbook_observed_sha256": before_hash,
        "normalized_response_sha256": normalized_hash,
        "canonical_sheet": manifest.canonical_sheet,
        "canonical_fields": list(manifest.fields),
        "unique_b1_case_count": unique_b1_count,
        "manifest_signed": manifest.is_signed,
    }
    return SourceAuthorityResolution(
        manifest=manifest,
        normalized_rows=rows,
        conflict_ledger=ledger,
        consistency_audit=consistency,
        gate=gate,
    )


def write_authority_audit(
    output_dir: str | Path,
    resolution: SourceAuthorityResolution,
) -> dict[str, Path]:
    """Write the complete detached authority evidence package."""

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    normalized_hash = str(resolution.gate["normalized_response_observed_sha256"])
    hash_path = destination / "normalized_unassisted_responses.sha256"
    hash_path.write_text(normalized_hash + "\n", encoding="ascii", newline="\n")
    return {
        "manifest": write_json(
            destination / "response_authority_manifest.json",
            resolution.manifest.to_dict(),
        ),
        "consistency_audit": write_json(
            destination / "workbook_consistency_audit.json",
            resolution.consistency_audit,
        ),
        "conflict_ledger": write_csv(
            destination / "workbook_conflict_ledger.csv",
            resolution.conflict_ledger,
        ),
        "normalized_responses": write_jsonl(
            destination / "normalized_unassisted_responses.jsonl",
            resolution.normalized_rows,
        ),
        "normalized_responses_sha256": hash_path,
        "gate": write_json(destination / "source_authority_gate.json", resolution.gate),
    }

"""Fail-closed reader for the completed CARDIA-X validation workbook."""

from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
import re
import unicodedata
from typing import Mapping

from tm_ecg.clinical_validation.field_policy import ALLOWED_UNASSISTED_FIELDS
from tm_ecg.clinical_validation.models import CaseIdentity, RawPhysicianResponse
from tm_ecg.clinical_validation.source_authority import (
    SourceAuthorityError,
    SourceAuthorityResolution,
    allows_unsigned_source_exploration,
    load_response_authority,
    resolve_unassisted_source,
    write_authority_audit,
)


EXPECTED_RESPONSE_HEADERS_V5 = (
    "case_id",
    "dataset_branch",
    "display_id",
    "image_filename",
    "image_absolute_path",
    "cardia_x_route",
    "primary_label_or_abstention",
    "activated_rule_ids",
    "uncertainty_flags",
    "safety_note",
    "unassisted_status",
    "assisted_status",
    "b2_audit_status",
    "overall_status",
    "primary_diagnosis",
    "diagnostic_confidence",
    "ecg_quality",
    "ambiguous",
    "requires_additional_information",
    "evidence_rr_irregularity",
    "evidence_absent_p_waves",
    "evidence_wide_qrs",
    "evidence_bundle_branch_pattern",
    "evidence_premature_beat",
    "evidence_paced_morphology",
    "evidence_st_deviation",
    "evidence_qtc_abnormality",
    "evidence_lead_noise",
    "evidence_other",
    "rationale",
    "diagnosis_changed",
    "assisted_confidence",
    "utility",
    "soundness",
    "safety",
    "clarity",
    "abstention_appropriate",
    "conflict_plausible",
    "missed_feature",
    "misleading_or_unsafe_feature",
    "final_comment",
    "p_wave_delineation_plausible",
    "qrs_delineation_plausible",
    "t_wave_delineation_plausible",
    "lead_quality_reasonable",
    "lead_quality_utility",
    "morphology_issue",
    "morphology_comment",
)

EXPECTED_RESPONSE_HEADERS_V6 = (
    *EXPECTED_RESPONSE_HEADERS_V5[:41],
    "unassisted_review_completion_time",
    "assisted_output_reveal_time",
    "blinding_deviation",
    "deviation_explanation",
    "reviewer_attestation_unassisted_first",
    "morphology_issue",
    "morphology_comment",
)

# Backward-compatible public name used by the synthetic workbook fixtures.
EXPECTED_RESPONSE_HEADERS = EXPECTED_RESPONSE_HEADERS_V5

CASE_SHEET_CELLS = {
    "primary_diagnosis": "M7",
    "diagnostic_confidence": "M8",
    "ecg_quality": "M9",
    "ambiguous": "M10",
    "requires_additional_information": "M11",
    "evidence_rr_irregularity": "M13",
    "evidence_absent_p_waves": "M14",
    "evidence_wide_qrs": "M15",
    "evidence_bundle_branch_pattern": "M16",
    "evidence_premature_beat": "M17",
    "evidence_paced_morphology": "M18",
    "evidence_st_deviation": "M19",
    "evidence_qtc_abnormality": "M20",
    "evidence_lead_noise": "M21",
    "evidence_other": "M22",
    "rationale": "M24",
}

B2_FIELDS = (
    "p_wave_delineation_plausible",
    "qrs_delineation_plausible",
    "t_wave_delineation_plausible",
    "lead_quality_reasonable",
    "lead_quality_utility",
    "morphology_issue",
    "morphology_comment",
)

CARDIA_X_FIELDS = (
    "case_id",
    "cardia_x_route",
    "primary_label_or_abstention",
    "activated_rule_ids",
    "uncertainty_flags",
    "safety_note",
)

ASSISTED_REVIEW_FIELDS = (
    "case_id",
    "diagnosis_changed",
    "assisted_confidence",
    "utility",
    "soundness",
    "safety",
    "clarity",
    "abstention_appropriate",
    "conflict_plausible",
    "missed_feature",
    "misleading_or_unsafe_feature",
    "final_comment",
)


class WorkbookSchemaError(ValueError):
    pass


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return re.sub(r"\s+", " ", text)


def _numeric(value: object) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(str(value))
    except (TypeError, ValueError) as exc:
        raise WorkbookSchemaError(f"Expected numeric workbook value, got {value!r}") from exc


def _response_export_schema(headers: tuple[str, ...]) -> str:
    if headers == EXPECTED_RESPONSE_HEADERS_V5:
        return "response_export_v5"
    if headers == EXPECTED_RESPONSE_HEADERS_V6:
        return "response_export_v6"
    expected_union = set(EXPECTED_RESPONSE_HEADERS_V5) | set(EXPECTED_RESPONSE_HEADERS_V6)
    missing = sorted(expected_union - set(headers))
    extra = sorted(set(headers) - expected_union)
    raise WorkbookSchemaError(f"Response_Export schema mismatch; missing={missing}, extra={extra}")


def load_provenance(path: str | Path) -> dict[str, dict[str, object]]:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    rows = payload.get("case_provenance")
    if not isinstance(rows, dict):
        raise WorkbookSchemaError("Provenance must contain a case_provenance object")
    return {str(key): dict(value) for key, value in rows.items()}


@dataclass(frozen=True, slots=True)
class WorkbookExtraction:
    identities: tuple[CaseIdentity, ...]
    physician_responses: tuple[RawPhysicianResponse, ...]
    b2_responses: tuple[dict[str, object], ...]
    cardia_x_outputs: tuple[dict[str, object], ...]
    assisted_review_responses: tuple[dict[str, object], ...]
    response_import_audit: dict[str, object]
    rule_review_summary: dict[str, object]


def _rule_review_summary(sheet: object) -> dict[str, object]:
    rows = list(sheet.iter_rows(min_row=5, max_row=15, values_only=True))  # type: ignore[attr-defined]
    if not rows:
        raise WorkbookSchemaError("Rule_Review has no header row")
    headers = [normalize_cell(value) for value in rows[0]]
    expected = {
        "scope",
        "rule_id",
        "antecedent combination",
        "consequent label",
        "A1 support",
        "A2 conservative",
        "A3 understandable",
        "A5 useful",
    }
    if not expected <= set(headers):
        raise WorkbookSchemaError("Rule_Review schema is incomplete")
    payload_rows = [
        {header: value for header, value in zip(headers, row, strict=False)}
        for row in rows[1:]
        if normalize_cell(row[0])
    ]
    score_fields = ("A1 support", "A2 conservative", "A3 understandable", "A5 useful")
    numeric_scores = [
        float(row[field])
        for row in payload_rows
        for field in score_fields
        if row.get(field) not in {None, ""}
    ]
    definitions_complete = sum(
        bool(normalize_cell(row.get("antecedent combination")))
        and bool(normalize_cell(row.get("consequent label")))
        for row in payload_rows
    )
    scored_rules = sum(
        all(row.get(field) not in {None, ""} for field in score_fields) for row in payload_rows
    )
    complete = len(payload_rows) == 10 and definitions_complete == 10 and scored_rules == 10
    return {
        "status": "complete" if complete else "not_estimable_incomplete_rulebook",
        "expected_rules": 10,
        "expected_b1_rules": 8,
        "expected_b2_rules": 2,
        "template_rows": len(payload_rows),
        "supplied_rule_ids": sum(bool(normalize_cell(row.get("rule_id"))) for row in payload_rows),
        "complete_rule_definitions": definitions_complete,
        "fully_scored_rules": scored_rules,
        "observed_likert_cells": len(numeric_scores),
        "mean_likert": sum(numeric_scores) / len(numeric_scores) if numeric_scores else None,
        "reason": (
            "Rule antecedents/consequents and physician scores were not supplied; "
            "rule soundness cannot be inferred from case-level rule IDs."
            if not complete
            else ""
        ),
    }


def _worksheet_rows(sheet: object) -> tuple[list[str], list[dict[str, object]]]:
    values = list(sheet.iter_rows(values_only=True))  # type: ignore[attr-defined]
    if not values:
        raise WorkbookSchemaError("Response_Export is empty")
    headers = [normalize_cell(value) for value in values[0]]
    rows = []
    for row in values[1:]:
        if not any(value not in {None, ""} for value in row):
            continue
        rows.append({header: value for header, value in zip(headers, row, strict=False)})
    return headers, rows


def _case_ids_from_index(sheet: object) -> list[str]:
    values = list(sheet.iter_rows(values_only=True))  # type: ignore[attr-defined]
    if not values or normalize_cell(values[0][0]) != "case_id":
        raise WorkbookSchemaError("Case_Index must begin with a case_id column")
    return [normalize_cell(row[0]) for row in values[1:] if normalize_cell(row[0])]


def _assert_case_sheet_reconciliation(
    workbook: object,
    case_id: str,
    response_row: Mapping[str, object],
) -> dict[str, str]:
    if case_id not in workbook.sheetnames:  # type: ignore[attr-defined]
        raise WorkbookSchemaError(f"Missing case worksheet {case_id}")
    sheet = workbook[case_id]  # type: ignore[index]
    if normalize_cell(sheet["M5"].value) != case_id:
        raise WorkbookSchemaError(f"Case identity mismatch on worksheet {case_id}")
    source_cells: dict[str, str] = {}
    for field, coordinate in CASE_SHEET_CELLS.items():
        sheet_value = normalize_cell(sheet[coordinate].value)
        export_value = normalize_cell(response_row.get(field))
        if sheet_value != export_value:
            raise WorkbookSchemaError(
                f"{case_id} field {field} disagrees between {coordinate} and Response_Export"
            )
        source_cells[field] = f"{case_id}!{coordinate}"
    return source_cells


def read_completed_workbook(
    workbook_path: str | Path,
    provenance_path: str | Path,
    *,
    reconciliation_mode: str = "strict_reconciliation",
    response_authority_path: str | Path | None = None,
    authority_output_dir: str | Path | None = None,
    allow_unsigned_source_exploration: bool = False,
) -> WorkbookExtraction:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Clinical validation requires openpyxl>=3.1") from exc

    workbook_source = Path(workbook_path)
    provenance = load_provenance(provenance_path)
    if reconciliation_mode not in {"strict_reconciliation", "authority_manifest"}:
        raise WorkbookSchemaError(
            "reconciliation_mode must be strict_reconciliation or authority_manifest"
        )
    if reconciliation_mode == "strict_reconciliation" and response_authority_path is not None:
        raise WorkbookSchemaError(
            "A response-authority manifest may only be used in authority_manifest mode"
        )
    if reconciliation_mode == "authority_manifest" and response_authority_path is None:
        raise WorkbookSchemaError(
            "authority_manifest mode requires an explicit response-authority manifest"
        )

    authority_resolution: SourceAuthorityResolution | None = None
    authority_rows: dict[str, Mapping[str, object]] = {}
    if reconciliation_mode == "authority_manifest":
        assert response_authority_path is not None
        try:
            authority_manifest = load_response_authority(response_authority_path)
            original_case_ids = {
                case_id: (normalize_cell(row.get("duplicate_of_case_id")) or case_id)
                for case_id, row in provenance.items()
            }
            authority_resolution = resolve_unassisted_source(
                workbook_source,
                authority_manifest,
                original_case_ids=original_case_ids,
            )
            if authority_output_dir is not None:
                write_authority_audit(authority_output_dir, authority_resolution)
            if not authority_resolution.gate["passed"] and not (
                allow_unsigned_source_exploration
                and allows_unsigned_source_exploration(authority_resolution.gate)
            ):
                failed = sorted(
                    key for key, passed in authority_resolution.gate["checks"].items() if not passed
                )
                raise WorkbookSchemaError(
                    f"Source-authority gate failed before response coding: {failed}"
                )
            authority_rows = {
                str(row["workbook_case_id"]): row for row in authority_resolution.normalized_rows
            }
        except SourceAuthorityError as exc:
            raise WorkbookSchemaError(str(exc)) from exc

    structure = load_workbook(workbook_source, read_only=True, data_only=False)
    values = load_workbook(workbook_source, read_only=True, data_only=True)
    required_sheets = {"Case_Index", "Response_Export", "Rule_Review"}
    missing_sheets = sorted(required_sheets - set(structure.sheetnames))
    if missing_sheets:
        raise WorkbookSchemaError(f"Missing workbook sheets: {missing_sheets}")

    headers, response_rows = _worksheet_rows(values["Response_Export"])
    response_export_schema = _response_export_schema(tuple(headers))
    index_case_ids = _case_ids_from_index(values["Case_Index"])
    response_case_ids = [normalize_cell(row["case_id"]) for row in response_rows]
    if index_case_ids != response_case_ids:
        raise WorkbookSchemaError("Case_Index and Response_Export case ordering differ")
    if len(response_case_ids) != 150 or len(set(response_case_ids)) != 150:
        raise WorkbookSchemaError("Expected exactly 150 distinct workbook case IDs")

    duplicate_targets = {
        normalize_cell(row.get("duplicate_of_case_id"))
        for row in provenance.values()
        if normalize_cell(row.get("duplicate_of_case_id"))
    }
    identities: list[CaseIdentity] = []
    responses: list[RawPhysicianResponse] = []
    b2_responses: list[dict[str, object]] = []
    cardia_x_outputs: list[dict[str, object]] = []
    assisted_review_responses: list[dict[str, object]] = []
    route_mismatches: list[str] = []
    for ordinal, row in enumerate(response_rows, start=2):
        case_id = normalize_cell(row["case_id"])
        provenance_row = provenance.get(case_id)
        if provenance_row is None:
            raise WorkbookSchemaError(f"Missing provenance for {case_id}")
        branch = normalize_cell(row["dataset_branch"])
        dataset = "ptbxl" if branch == "B1" else "ludb" if branch == "B2" else ""
        if not dataset:
            raise WorkbookSchemaError(f"Unknown dataset branch {branch!r} for {case_id}")
        duplicate_of = normalize_cell(provenance_row.get("duplicate_of_case_id"))
        original_case_id = duplicate_of or case_id
        route = normalize_cell(provenance_row.get("route_category"))
        if route != normalize_cell(row["cardia_x_route"]):
            route_mismatches.append(case_id)
        identity = CaseIdentity(
            workbook_case_id=case_id,
            original_case_id=original_case_id,
            dataset=dataset,
            source_record_id=normalize_cell(provenance_row.get("record_id")),
            route=route,
            repeat_group_id=original_case_id
            if (duplicate_of or case_id in duplicate_targets)
            else None,
            row_ordinal=ordinal,
        )
        identities.append(identity)

        if reconciliation_mode == "strict_reconciliation":
            source_cells = _assert_case_sheet_reconciliation(values, case_id, row)
        else:
            authority_row = authority_rows.get(case_id)
            if authority_row is None:
                raise WorkbookSchemaError(f"Canonical response authority is missing {case_id}")
            raw_source_cells = authority_row.get("source_cells")
            if not isinstance(raw_source_cells, Mapping):
                raise WorkbookSchemaError(
                    f"Canonical response authority has no source cells for {case_id}"
                )
            source_cells = {field: str(raw_source_cells[field]) for field in CASE_SHEET_CELLS}
        evidence = {
            field: normalize_cell(row.get(field))
            for field in ALLOWED_UNASSISTED_FIELDS
            if field.startswith("evidence_")
        }
        response = RawPhysicianResponse.from_dict(
            {
                "case_id": case_id,
                "primary_diagnosis": normalize_cell(row.get("primary_diagnosis")),
                "rationale": normalize_cell(row.get("rationale")),
                "diagnostic_confidence": _numeric(row.get("diagnostic_confidence")),
                "ecg_quality": _numeric(row.get("ecg_quality")),
                "ambiguous": normalize_cell(row.get("ambiguous")),
                "requires_additional_information": normalize_cell(
                    row.get("requires_additional_information")
                ),
                "evidence": evidence,
                "source_cells": source_cells,
            }
        )
        responses.append(response)
        cardia_x_outputs.append({field: row.get(field) for field in CARDIA_X_FIELDS})
        assisted_review_responses.append(
            {
                **{field: row.get(field) for field in ASSISTED_REVIEW_FIELDS},
                "dataset_branch": branch,
                "route": route,
            }
        )
        if dataset == "ludb":
            b2_responses.append(
                {
                    "case_id": case_id,
                    **{field: row.get(field) for field in B2_FIELDS},
                    "source_cells": {
                        field: (
                            f"Response_Export!{field}"
                            if field in row
                            else "absent_from_Response_Export"
                        )
                        for field in B2_FIELDS
                    },
                }
            )

    b1 = [item for item in identities if item.dataset == "ptbxl"]
    b2 = [item for item in identities if item.dataset == "ludb"]
    b1_unique = [item for item in b1 if not item.is_repeat]
    route_counts: dict[str, int] = {}
    for item in b1_unique:
        route_counts[item.route] = route_counts.get(item.route, 0) + 1
    expected_routes = {"exact_match": 30, "conflict_region": 30, "structural_abstention": 40}
    if len(b1) != 110 or len(b1_unique) != 100 or len(b2) != 40:
        raise WorkbookSchemaError(
            f"Population mismatch: B1={len(b1)}, unique B1={len(b1_unique)}, B2={len(b2)}"
        )
    if route_counts != expected_routes:
        raise WorkbookSchemaError(f"Unique-B1 route mismatch: {route_counts}")
    if route_mismatches:
        raise WorkbookSchemaError(f"Workbook/provenance route mismatches: {route_mismatches}")

    audit = {
        "workbook_sheets": len(values.sheetnames),
        "total_rows": len(identities),
        "b1_rows": len(b1),
        "b1_unique": len(b1_unique),
        "b1_repeats": sum(item.is_repeat for item in b1),
        "b2_rows": len(b2),
        "unique_b1_route_counts": route_counts,
        "case_index_reconciled": True,
        "case_sheet_values_reconciled": reconciliation_mode == "strict_reconciliation",
        "reconciliation_mode": reconciliation_mode,
        "canonical_unassisted_response_source": (
            "case_sheet_reconciled_with_Response_Export"
            if reconciliation_mode == "strict_reconciliation"
            else "Response_Export"
        ),
        "response_headers": headers,
        "response_export_schema": response_export_schema,
        "b2_morphology_source": (
            "Response_Export_partial; structured B2 ratings unavailable"
            if response_export_schema == "response_export_v6"
            else "Response_Export"
        ),
        "b2_missing_response_export_fields": (
            list(B2_FIELDS[:5]) if response_export_schema == "response_export_v6" else []
        ),
    }
    if authority_resolution is not None:
        audit["source_authority_gate"] = authority_resolution.gate
        audit["workbook_consistency_audit"] = authority_resolution.consistency_audit
    rule_review = _rule_review_summary(values["Rule_Review"])
    structure.close()
    values.close()
    return WorkbookExtraction(
        tuple(identities),
        tuple(responses),
        tuple(b2_responses),
        tuple(cardia_x_outputs),
        tuple(assisted_review_responses),
        audit,
        rule_review,
    )
